import xlsxwriter
import os
import sys
import pytest
import datetime
import openpyxl
# Appending current directory to path allows importing for 'pytest' command line
sys.path.append('.')
import alignment
from TestData import test_wb_dict1, test_wb_dict2


@pytest.fixture(scope="module")
def files():
    # Bad empty files
    f1 = open("text_file.txt", "w+")
    f2 = open("~$temp_file.xlsx", "w+")
    f3 = open("01-23-45_date1.xlsx", "w+")
    f4 = open("2018-3-12_date2.xlsx", "w+")
    f1.close()
    f2.close()
    f3.close()
    f4.close()

    # Useful files
    row_1 = ['random_thing_top', 'HomoSapiens', 'c:/file/path/thing.gz']
    row_2 = ['read count', 'Percentage...', 'nucleotide seq', 'nucleotide2',
             'min quality', 'CDR3 amino acid sequence', 'other thing']
    column_A1 = [x for x in range(20)]
    column_A2 = [100-x for x in range(20)]
    column_B = [1/((100+x)**2) for x in range(20)]
    column_C = ['ABCDE' for x in range(20)]
    column_D = ['FGHI' for x in range(20)]
    column_E = [35 for x in range(20)]
    column_F1 = ['THE', 'CAT', 'IN', 'THE', 'HAT',
                'CCCDDDEE', 'FGHIKAA', 'OOPIS~JA', 'JKA', 'BDEDF',
                'SAT', 'ATOP', 'THE', 'VAT', 'OF',
                'SKAT', 'HOW', 'CAT', 'LIKE', 'CXCJKJ']
    column_F2 = ['ABCDEFG', 'HIGJKLM', 'OPQRS~TU', 'VWXYZ', 'AAAABBBB',
                'CCCDDDEE', 'FGHIKAA', 'OOPIS~JA', 'JKA', 'BDEDF',
                'EAJJKE', 'JKJLAJEL', 'EUWOMJS', 'WJIS~JSO', 'SEUIS',
                'IEJUDU', 'SUIOUS', 'AJIFJF', 'JDUIOS', 'CXCJKJ']
    column_G = ['TRAVE13-blah' for x in range(20)]


    top_rows = [row_1, row_2]
    data_columns1 = [column_A1, column_B, column_C, column_D,
            column_E, column_F1, column_G]
    data_columns2 = [column_A2, column_B, column_C, column_D,
            column_E, column_F2, column_G]

    wb1 = xlsxwriter.Workbook('test_book_1.xlsx')
    ws1 = wb1.add_worksheet('test_worksheet_1')

    wb2 = xlsxwriter.Workbook('test_book_2.xlsx')
    ws2 = wb2.add_worksheet('test_worksheet_2')

    def fill_excel(ws, data_columns):

        # Write headers
        row = 0
        col = 0
        for top_row in top_rows:
            for item in top_row:
                ws.write(row, col, item)
                col += 1
            col = 0
            row +=1

        # Write data
        row = 2
        col = 0
        for column in data_columns:
            for item in column:
                ws.write(row, col, item)
                row += 1
            row = 2
            col += 1

    fill_excel(ws1, data_columns1)
    fill_excel(ws2, data_columns2)

    wb1.close()
    wb2.close()

    yield

    new_file_name, directory_path, current_file_path = directory_info()
    delete_list = ["text_file.txt", "~$temp_file.xlsx",
                         "01-23-45_date1.xlsx", "2018-3-12_date2.xlsx",
                         "test_book_1.xlsx", "test_book_2.xlsx", new_file_name]
    delete_files(delete_list)


def delete_files(delete_list):
    new_file_name, directory_path, current_file_path = directory_info()
    for file in os.listdir(directory_path):
        filename = os.fsdecode(file)
        current_file_path = os.path.join(directory_path, filename)
        if filename in delete_list:
            os.remove(current_file_path)


def test_get_file_count(files):
    file_count, file_list, header_list = alignment.get_file_count()
    assert file_count == 2
    assert header_list == ['test_book_1', 'test_book_2']
    assert file_list == ["test_book_1.xlsx", "test_book_2.xlsx"]


def test_files_iterated_over():
    file_list = ["test_book_1.xlsx", "test_book_2.xlsx"]
    wb_dict = alignment.iterate_over_files(2, file_list)
    assert wb_dict == test_wb_dict1


def directory_info():
    directory_path = os.path.dirname(os.path.realpath(__file__))
    today_date = datetime.datetime.date(datetime.datetime.now())
    new_file_name = "1test_" + str(today_date) + "_combined_matrix.xlsx"
    current_file_path = os.path.join(directory_path, new_file_name,)

    return new_file_name, directory_path, current_file_path


@pytest.mark.parametrize("test_input1, test_input2, exp_out1, exp_out2, exp_out3",
                         [
                             (test_wb_dict1, ['Motif', 'head1', 'head2'],
                              'THE', 82, None),
                             (test_wb_dict2, ['header'+str(x) for x in range(20)],
                              'THE', None, 2),
                         ])
def test_write_to_workbook(test_input1, test_input2,
                           exp_out1, exp_out2, exp_out3):

    new_file_name, directory_path, current_file_path = directory_info()

    alignment.write_to_workbook(test_input1, test_input2, is_test=True)
    assert new_file_name in os.listdir(directory_path)

    wb = openpyxl.load_workbook(current_file_path)
    ws = wb.active

    assert ws['A2'].value == exp_out1
    assert ws['C32'].value == exp_out2
    assert ws['U8'].value == exp_out3

    delete_files([new_file_name])


def test_alignment():
    file_count, file_list, header_list = alignment.get_file_count()
    wb_dict = alignment.iterate_over_files(file_count, file_list)
    alignment.write_to_workbook(wb_dict, header_list, is_test=True)

    new_file_name, directory_path, current_file_path = directory_info()
    assert new_file_name in os.listdir(directory_path)

    wb = openpyxl.load_workbook(current_file_path)
    ws = wb.active

    assert ws.title == "combined_matrix"

    assert ws['A1'].value == 'Motifs'
    assert ws['B1'].value == 'test_book_1'
    assert ws['A2'].value == 'THE'
    assert ws['A6'].value == 'CCCDDDEE'
    assert ws['B6'].value == 5
    assert ws['C32'].value == 82
    assert ws['U8'].value is None



if __name__ == '__main__':
    pass
    files()
