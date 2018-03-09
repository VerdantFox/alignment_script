# https://openpyxl.readthedocs.io/en/stable/
from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active


# Mock data
file_count = 3



# Iterate through rows of current worksheet
for row in ws.iter_rows(
        min_row=2, min_col=1, max_col=file_count + 1):
    aa_seq = None
    aa_seq_count = None

    for cell in row:
        if cell.column == 'A':
            # Assign amino acid sequence variable for current row
            aa_seq_count = cell.value
        elif cell.column == 'F':
            # Assign sequence count variable for current row
            aa_seq = cell.value