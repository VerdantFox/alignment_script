# https://openpyxl.readthedocs.io/en/stable/
from openpyxl import load_workbook
# https://xlsxwriter.readthedocs.io/
import xlsxwriter
import os
import datetime


def get_file_count():
    """Gets the number of files to read and creates list from them"""

    file_list = []
    file_count = 0
    # https://stackoverflow.com/questions/10377998/how-can-i-iterate-over-files-in-a-given-directory
    # This gives the directory path from which the .py file is being run
    directory_path = os.path.dirname(os.path.realpath(__file__))
    for file in os.listdir(directory_path):
        filename = os.fsdecode(file)

        if filename.endswith(".xlsx") and not filename.startswith('~'):
            file_count += 1
            file_list.append(filename)

    print()
    print(f"file count is {file_count}")
    print(f"files working with:")
    print(file_list)
    print()

    return file_count, file_list, directory_path


def get_current_wb_headers(ws, header_lists):
    """Counts current workbook's headers and adds them to header_list"""

    header_count = 0
    current_header_list = []

    # Iterate through rows of current worksheet
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=2):
        for cell in row:
            header_count += 1
            current_header_list.append(cell.value)

    header_lists.append(current_header_list)

    return header_count


def open_workbooks(file_list, directory_path):
    """Opens workbooks, gets count of and list of headers"""

    print("******** GATHERING HEADERS ************")
    header_lists = []
    header_count = 0

    wb_list = []

    # Iterate over each file in the current folder
    for filename in file_list:
        # Open current workbook and go to its worksheet for reading
        print(f"\n******* Opening {filename} *******")
        current_file_path = os.path.join(directory_path, filename)
        current_wb = load_workbook(current_file_path)
        wb_list.append(current_wb)
        current_ws = current_wb.active
        print(f"\n******** getting headers from {filename} ********")

        current_header_count = get_current_wb_headers(
            current_ws, header_lists)

        # print(header_count)
        # print(current_header_count)
        header_count += current_header_count

    return header_count, header_lists, wb_list


def iterate_over_workbooks(header_count, wb_list, header_lists, file_list):
    """Reads through each excel file, adds sequences and counts to dict"""

    combo_dict = dict()
    starting_dict_col = 0

    # Iterate over each file in the current folder
    for index, current_wb in enumerate(wb_list):

        current_ws = current_wb.active

        print(f"\n**** Adding {file_list[index]} to dictionary******")
        # Iterate through rows of current worksheet
        for row in current_ws.iter_rows(
                min_row=2, min_col=1):
            dict_col = starting_dict_col
            aa_seq = None

            for cell in row:
                if cell.column == 'A':
                    # Assign sequence count variable for current row
                    aa_seq = cell.value

                    # Add aa_seq to dict, populate all columns 0
                    if aa_seq not in combo_dict:
                        combo_dict[aa_seq] = [0 for x in range(header_count)]
                else:
                    if cell.value is not None:
                        # Add amino acid sequence count to current column of dict
                        combo_dict[aa_seq][dict_col] += cell.value

                    # Move to next column of dict
                    dict_col += 1

        current_header_count = len(header_lists[index])
        starting_dict_col += current_header_count

    return combo_dict


def write_to_workbook(wb_dict, header_lists, is_test=False):
    """Writes a .xlsx excel workbook from the dictionary given and saves"""

    header_list = []
    for hlist in header_lists:
        header_list.extend(hlist)

    # Get today's date for naming purposes
    today_date = datetime.datetime.date(datetime.datetime.now())
    if is_test:
        new_file_name = "1test_" + str(today_date) + "_combined_matrix.xlsx"
    else:
        new_file_name = str(today_date) + "_multi_matrix.xlsx"

    # Name excel workbook and excel worksheet
    workbook = xlsxwriter.Workbook(new_file_name)
    worksheet = workbook.add_worksheet("multi_matrix")

    # Start from the first cell. Rows and columns are zero indexed.
    row = 0
    col = 0

    print("\n******* WRITING EXCEL WORKBOOK *********\n")
    # Add all headers to excel file
    worksheet.write(row, col, 'Motifs')
    col += 1
    for header in header_list:
        worksheet.write(row, col, header)
        col += 1

    # write dictionary to excel file
    for aa_seq in wb_dict:
        col = 0
        row += 1
        worksheet.write(row, col, aa_seq)
        for aa_seq_count in wb_dict[aa_seq]:
            col += 1
            worksheet.write(row, col, aa_seq_count)

    # Close (and thus save) excel workbook
    workbook.close()

    print("\n************** FILE SAVED **************\n")


if __name__ == '__main__':
    file_count, file_list, directory_path = get_file_count()
    header_count, header_lists, wb_list = open_workbooks(
        file_list, directory_path)
    starting_dict_col = 0
    combo_dict = iterate_over_workbooks(
        header_count, wb_list, header_lists, file_list)
    write_to_workbook(combo_dict, header_lists)
