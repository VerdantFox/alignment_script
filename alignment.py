# https://openpyxl.readthedocs.io/en/stable/
from openpyxl import load_workbook
# https://xlsxwriter.readthedocs.io/
import xlsxwriter
import os
import datetime


def get_file_count():
    """Gets the number of files to read and creates list from them"""

    header_list = []
    file_count = 0
    directory_path = os.path.dirname(os.path.realpath(__file__))
    for file in os.listdir(directory_path):
        filename = os.fsdecode(file)

        if filename.endswith(".xlsx") and not filename.startswith(
                ('~', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9')):
            file_count += 1
            header_list.append(filename.strip('.xlsx'))

    print()
    print(f"file count is {file_count}")
    print(f"files working with:")
    print(header_list)
    print()

    return file_count, header_list


def iterate_over_files(file_count):
    """Reads through each excel file, adds sequences and counts to dict"""

    column_counter = 0
    wb_dict = dict()

    # https://stackoverflow.com/questions/10377998/how-can-i-iterate-over-files-in-a-given-directory
    # This gives the directory path from which the .py file is being run
    directory_path = os.path.dirname(os.path.realpath(__file__))

    # Iterate over each file in the current folder
    for file in os.listdir(directory_path):
        filename = os.fsdecode(file)
        # Specify only excel files
        # Exclude files that start with a number (our exit file, a date does so)
        # Exclude files that start with a '~' (these are temp, opened files)
        if filename.endswith(".xlsx") and not filename.startswith(
                ('~', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9')):

            # Open current workbook and go to its worksheet for reading
            current_file_path = os.path.join(directory_path, filename)
            print(f"\n******* Opening {filename} *******")
            current_wb = load_workbook(current_file_path)
            current_ws = current_wb.active
            print(f"\n******** working on {current_wb.sheetnames[0]} ********")

            # Iterate through rows of current worksheet
            for row in current_ws.iter_rows(min_row=3, min_col=1, max_col=6):
                aa_seq = None
                aa_seq_count = None

                for cell in row:
                    if cell.column == 'A':
                        # Assign amino acid sequence variable for current row
                        aa_seq_count = cell.value
                    elif cell.column == 'F':
                        # Assign sequence count variable for current row
                        aa_seq = cell.value

                # Add amino acid sequence to workbook and populate all columns 0
                if aa_seq not in wb_dict:
                    wb_dict[aa_seq] = [0 for x in range(file_count)]

                # Add amino acid sequence count to current column of dict
                wb_dict[aa_seq][column_counter] += aa_seq_count

            # Move to next column
            column_counter += 1

    return wb_dict


def write_to_workbook(wb_dict, header_list, is_test=False):
    """Writes a .xlsx excel workbook from the dictionary given and saves"""

    # Get today's date for naming purposes
    today_date = datetime.datetime.date(datetime.datetime.now())
    if is_test:
        new_file_name = "1test_" + str(today_date) + "_combined_matrix.xlsx"
    else:
        new_file_name = str(today_date) + "_combined_matrix.xlsx"

    # Name excel workbook and excel worksheet
    workbook = xlsxwriter.Workbook(new_file_name)
    worksheet = workbook.add_worksheet("combined_matrix")

    # Start from the first cell. Rows and columns are zero indexed.
    row = 0
    col = 0

    print("\n******* WRITING EXCEL WORKBOOK *********\n")
    # Add all headers to excel file
    worksheet.write(row, col, 'Motifs')
    col += 1
    for header in header_list:
        worksheet.write(row, col, header)
        col += 1

    # write dictionary to excel file
    for aa_seq in wb_dict:
        col = 0
        row += 1
        worksheet.write(row, col, aa_seq)
        for aa_seq_count in wb_dict[aa_seq]:
            col += 1
            worksheet.write(row, col, aa_seq_count)

    # Close (and thus save) excel workbook
    workbook.close()

    print("\n************** FILE SAVED **************\n")


if __name__ == '__main__':

    file_count, header_list = get_file_count()

    wb_dict = iterate_over_files(file_count)

    write_to_workbook(wb_dict, header_list)
